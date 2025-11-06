#!/usr/bin/env python3
"""
TSN Conformance Test Report Generator
Generates professional test reports in HTML, PDF, JSON, and Excel formats
"""

import json
import argparse
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any
import sys

try:
    import pandas as pd
    import matplotlib.pyplot as plt
    import matplotlib
    matplotlib.use('Agg')  # Non-interactive backend
except ImportError:
    print("Warning: pandas and/or matplotlib not installed. Install with:")
    print("  pip3 install --user pandas matplotlib")
    sys.exit(1)


class ConformanceReportGenerator:
    """Generate conformance test reports in various formats"""

    def __init__(self, results_file: str):
        """Initialize report generator

        Args:
            results_file: Path to JSON test results file
        """
        self.results_file = Path(results_file)
        with open(self.results_file) as f:
            self.data = json.load(f)

        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def generate_html(self, output_file: str) -> None:
        """Generate HTML report"""
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>TSN Conformance Test Report</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
            background: #f5f5f5;
        }}
        .header {{
            background: #0066CC;
            color: white;
            padding: 30px;
            border-radius: 8px;
            margin-bottom: 20px;
        }}
        .header h1 {{
            margin: 0;
            font-size: 32px;
        }}
        .header p {{
            margin: 10px 0 0 0;
            opacity: 0.9;
        }}
        .summary {{
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 15px;
            margin-bottom: 20px;
        }}
        .summary-card {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .summary-card h3 {{
            margin: 0;
            color: #666;
            font-size: 14px;
            font-weight: normal;
        }}
        .summary-card p {{
            margin: 10px 0 0 0;
            font-size: 32px;
            font-weight: bold;
            color: #0066CC;
        }}
        .test-results {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        th, td {{
            text-align: left;
            padding: 12px;
            border-bottom: 1px solid #e0e0e0;
        }}
        th {{
            background: #f8f9fa;
            font-weight: 600;
        }}
        .pass {{ color: #0066CC; font-weight: bold; }}
        .fail {{ color: #DC3545; font-weight: bold; }}
        .metric {{
            display: inline-block;
            padding: 4px 8px;
            background: #e3f2fd;
            border-radius: 4px;
            font-family: monospace;
            margin: 2px;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>TSN Conformance Test Report</h1>
        <p>Generated: {self.timestamp}</p>
        <p>Test Suite: {self.data.get('test_suite', 'N/A')}</p>
    </div>

    <div class="summary">
        <div class="summary-card">
            <h3>Total Tests</h3>
            <p>{self.data.get('total_tests', 0)}</p>
        </div>
        <div class="summary-card">
            <h3>Passed</h3>
            <p class="pass">{self.data.get('passed', 0)}</p>
        </div>
        <div class="summary-card">
            <h3>Failed</h3>
            <p class="fail">{self.data.get('failed', 0)}</p>
        </div>
        <div class="summary-card">
            <h3>Pass Rate</h3>
            <p>{self._calculate_pass_rate():.1f}%</p>
        </div>
    </div>

    <div class="test-results">
        <h2>Test Results</h2>
        {self._generate_results_table()}
    </div>
</body>
</html>"""

        with open(output_file, 'w') as f:
            f.write(html)
        print(f"HTML report generated: {output_file}")

    def generate_json(self, output_file: str) -> None:
        """Generate JSON report"""
        report = {
            "timestamp": self.timestamp,
            "summary": {
                "total_tests": self.data.get('total_tests', 0),
                "passed": self.data.get('passed', 0),
                "failed": self.data.get('failed', 0),
                "pass_rate": self._calculate_pass_rate()
            },
            "tests": self.data.get('tests', []),
            "system_info": self.data.get('system_info', {})
        }

        with open(output_file, 'w') as f:
            json.dump(report, f, indent=2)
        print(f"JSON report generated: {output_file}")

    def generate_excel(self, output_file: str) -> None:
        """Generate Excel report"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("Error: openpyxl not installed. Install with:")
            print("  pip3 install --user openpyxl")
            return

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"

        # Header
        headers = ["Test Name", "Standard", "Status", "Measured", "Required", "Result"]
        ws.append(headers)

        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Add test results
        for test in self.data.get('tests', []):
            ws.append([
                test.get('name', ''),
                test.get('standard', ''),
                test.get('status', ''),
                test.get('measured', ''),
                test.get('required', ''),
                test.get('result', '')
            ])

        # Save workbook
        wb.save(output_file)
        print(f"Excel report generated: {output_file}")

    def _calculate_pass_rate(self) -> float:
        """Calculate pass rate percentage"""
        total = self.data.get('total_tests', 0)
        if total == 0:
            return 0.0
        passed = self.data.get('passed', 0)
        return (passed / total) * 100

    def _generate_results_table(self) -> str:
        """Generate HTML table of test results"""
        html = "<table>\n<tr><th>Test Name</th><th>Standard</th><th>Status</th><th>Measured</th><th>Required</th><th>Result</th></tr>\n"

        for test in self.data.get('tests', []):
            status_class = "pass" if test.get('status') == 'PASS' else "fail"
            html += f"""<tr>
    <td>{test.get('name', '')}</td>
    <td><span class="metric">{test.get('standard', '')}</span></td>
    <td class="{status_class}">{test.get('status', '')}</td>
    <td>{test.get('measured', '')}</td>
    <td>{test.get('required', '')}</td>
    <td>{test.get('result', '')}</td>
</tr>\n"""

        html += "</table>"
        return html


def main():
    parser = argparse.ArgumentParser(description='Generate TSN conformance test reports')
    parser.add_argument('--results', required=True, help='Path to JSON results file')
    parser.add_argument('--format', choices=['html', 'json', 'excel', 'all'], default='html',
                        help='Report format (default: html)')
    parser.add_argument('--output', help='Output file path (default: auto-generated)')

    args = parser.parse_args()

    if not Path(args.results).exists():
        print(f"Error: Results file not found: {args.results}")
        sys.exit(1)

    generator = ConformanceReportGenerator(args.results)

    # Generate base output name
    if args.output:
        output_base = Path(args.output).stem
        output_dir = Path(args.output).parent
    else:
        output_base = f"conformance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        output_dir = Path('results')
        output_dir.mkdir(exist_ok=True)

    # Generate requested formats
    if args.format == 'html' or args.format == 'all':
        output_file = output_dir / f"{output_base}.html"
        generator.generate_html(str(output_file))

    if args.format == 'json' or args.format == 'all':
        output_file = output_dir / f"{output_base}.json"
        generator.generate_json(str(output_file))

    if args.format == 'excel' or args.format == 'all':
        output_file = output_dir / f"{output_base}.xlsx"
        generator.generate_excel(str(output_file))


if __name__ == '__main__':
    main()
